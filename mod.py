import openpyxl as opx
import numpy as np
import pandas as pd
import matplotlib.pyplot as mpl
import sys

#"Constants", the cursor cells
FP_0 = 'B5'
FP_1 = 'B6'
SA_0 = 'B9'
SA_1 = 'B10'
SA_2 = 'B11'
SA_3 = 'B12'
SA_4 = 'B13'
SA_5 = 'B14'

def Update(fn):
    wb = opx.load_workbook(fn)
    ws = wb.active

    ws[FP_0] = 'New Value'
    ws[FP_1] = 'New Value'

    ws[SA_0] = 'New Value'
    ws[SA_1] = 'New Value'
    ws[SA_2] = 'New Value'
    ws[SA_3] = 'New Value'
    ws[SA_4] = 'New Value'
    ws[SA_5] = 'New Value'

    wb.save(fn)


def Find_Fingers(data):
    fingers = {}
    #TODO

    return fingers


def To_DataFrame(fn):
    df = pd.read_csv(fn, delimiter='\t')
    df.columns = ['y1', 'y2', 'y3', 'y4', 'y5', 'y6', 'y7', 'Volts'] #need to fingure this out still
    return df


def Plot_All(fingers, data_df):
    for col in data_df.drop('Volts', axis=1).columns:
        mpl.plot(data_df['Volts'], data_df[col], label=col)
    mpl.legend()
    mpl.show()


def main():
    if(len(sys.argv) != 3):
        print("ERROR: args must be of format \" [data_file.csv] [tuning_file.xlsx]\"")
        sys.exit()
    data_df = To_DataFrame(sys.argv[1])
    tuning_fn = sys.argv[2]

    fingers = Find_Fingers(data_df)
    Plot_All(fingers, data_df)
    


if __name__ == "__main__":
    main()
